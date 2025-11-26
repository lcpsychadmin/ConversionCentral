from __future__ import annotations

import csv
import json
import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional, Sequence
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, Response, UploadFile, status
from sqlalchemy import Column, func, text
from sqlalchemy.engine import Connection, Engine
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.sql import sqltypes

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.constants.audit_fields import AUDIT_FIELD_NAME_SET
from app.database import get_db
from app.ingestion.engine import get_ingestion_connection_params, get_ingestion_engine
from app.models import (
    DataDefinition,
    DataDefinitionField,
    DataDefinitionTable,
    DataObject,
    DataObjectSystem,
    ConstructedTable,
    Field,
    ProcessArea,
    System,
    Table,
)
from app.routers.data_definition import _ensure_audit_fields_for_definition_table
from app.schemas.entities import DataWarehouseTarget
from pydantic import ValidationError

from app.schemas.upload_data import (
    UploadDataColumn,
    UploadDataColumnOverride,
    UploadDataCreateResponse,
    UploadDataPreviewResponse,
    UploadTableMode,
)
from app.services.data_construction_sync import sync_construction_tables_for_definition
from app.services.constructed_data_store import ConstructedDataStore
from app.services.constructed_data_warehouse import (
    ConstructedDataWarehouse,
    ConstructedDataWarehouseError,
)
from app.services.ingestion_loader import SparkTableLoader, build_loader_plan


logger = logging.getLogger(__name__)

router = APIRouter(prefix="/upload-data", tags=["Upload Data"])

MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
SAMPLE_ROW_LIMIT = 50
IDENTIFIER_SANITIZE_PATTERN = re.compile(r"[^0-9a-zA-Z_]")
TRUE_VALUES = {"true", "t", "yes", "y", "1"}
FALSE_VALUES = {"false", "f", "no", "n", "0"}


@dataclass(frozen=True)
class _UploadMetadataResult:
    table_id: UUID
    data_definition_id: UUID
    data_definition_table_id: UUID
    constructed_table_id: Optional[UUID]


@router.post("/preview", response_model=UploadDataPreviewResponse)
async def preview_upload_data(
    file: UploadFile = File(...),
    has_header: bool = Form(True),
    delimiter: Optional[str] = Form(None),
) -> UploadDataPreviewResponse:
    data = await _read_upload_bytes(file)
    columns, rows = _parse_uploaded_table_data(
        data=data,
        filename=file.filename,
        has_header=has_header,
        delimiter=delimiter,
    )
    await file.close()
    sample_rows = rows[:SAMPLE_ROW_LIMIT]
    return UploadDataPreviewResponse(columns=columns, sample_rows=sample_rows, total_rows=len(rows))


@router.post("/create-table", response_model=UploadDataCreateResponse)
async def create_table_from_upload(
    request: Request,
    table_name: str = Form(...),
    product_team_id: UUID = Form(...),
    data_object_id: UUID = Form(...),
    system_id: UUID = Form(...),
    file: UploadFile = File(...),
    has_header: bool = Form(True),
    delimiter: Optional[str] = Form(None),
    target: DataWarehouseTarget = Form(DataWarehouseTarget.DATABRICKS_SQL),
    schema_name: Optional[str] = Form(None),
    catalog: Optional[str] = Form(None),
    mode: UploadTableMode = Form(UploadTableMode.CREATE),
    column_overrides: Optional[str] = Form(None),
    db: Session = Depends(get_db),
) -> UploadDataCreateResponse:
    # SAP HANA is not a supported warehouse target; only Databricks is allowed.
    # If the frontend were to supply a SAP HANA target it will be rejected by schema validation now.

    overrides_raw = column_overrides
    if overrides_raw is None:
        try:
            form_data = await request.form()
        except Exception:
            form_data = None
        else:
            candidate = form_data.get("column_overrides") if form_data is not None else None
            if candidate is not None:
                if isinstance(candidate, UploadFile):
                    raw_bytes = await candidate.read()
                    overrides_raw = raw_bytes.decode("utf-8")
                else:
                    overrides_raw = str(candidate)

    sanitized_table_name = _sanitize_table_name(table_name)
    data = await _read_upload_bytes(file)
    columns, rows = _parse_uploaded_table_data(
        data=data,
        filename=file.filename,
        has_header=has_header,
        delimiter=delimiter,
    )
    included_indexes = list(range(len(columns)))
    if overrides_raw:
        overrides = _parse_column_overrides(overrides_raw)
        if overrides:
            columns, included_indexes = _apply_column_overrides(columns, overrides)
    await file.close()
    rows = _filter_row_values(rows, included_indexes)

    engine = get_ingestion_engine()
    dialect_name = engine.dialect.name
    resolved_schema = _resolve_default_schema(target, schema_name, dialect_name)
    effective_catalog = (catalog.strip() if isinstance(catalog, str) else None) or None
    if dialect_name != "databricks":
        effective_catalog = None

    qualified_table = _build_qualified_table_name(
        sanitized_table_name,
        resolved_schema,
        effective_catalog if dialect_name == "databricks" else None,
        dialect_name,
    )

    logger.info(
        "upload:create-table:init name=%s rows=%d columns=%d overrides=%d dialect=%s",
        sanitized_table_name,
        len(rows),
        len(columns),
        len(overrides) if overrides_raw else 0,
        dialect_name,
    )

    try:
        with engine.begin() as connection:
            _apply_create_table(
                connection=connection,
                qualified_table=qualified_table,
                table_name=sanitized_table_name,
                schema_name=resolved_schema,
                columns=columns,
                rows=rows,
                mode=mode,
                target=target,
                dialect_name=dialect_name,
            )
        logger.info("upload:create-table:warehouse-created name=%s", sanitized_table_name)
    except SQLAlchemyError as exc:  # pragma: no cover - requires live warehouse dialect for full coverage
        message = str(getattr(exc, "orig", exc)) or str(exc)
        lowered = message.lower()
        if "exists" in lowered or "already" in lowered:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Table already exists.") from exc
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Unable to create table: {message}") from exc

    metadata: Optional[_UploadMetadataResult] = None
    try:
        metadata = _register_uploaded_table_metadata(
            table_name=sanitized_table_name,
            schema_name=resolved_schema,
            columns=columns,
            product_team_id=product_team_id,
            data_object_id=data_object_id,
            system_id=system_id,
            session=db,
        )
        if metadata and metadata.constructed_table_id and rows:
            _sync_constructed_table_rows(
                constructed_table_id=metadata.constructed_table_id,
                columns=columns,
                rows=rows,
                session=db,
                force_recreate_warehouse=(mode is UploadTableMode.REPLACE),
            )
        db.commit()
        logger.info(
            "upload:create-table:metadata-applied name=%s constructed=%s",
            sanitized_table_name,
            metadata.constructed_table_id if metadata else None,
        )
    except HTTPException:
        db.rollback()
        _drop_table_if_exists(engine, qualified_table)
        raise
    except Exception as exc:  # pragma: no cover - defensive fallback
        db.rollback()
        logger.exception("Unexpected error registering uploaded table metadata for '%s'", sanitized_table_name)
        _drop_table_if_exists(engine, qualified_table)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Unable to register uploaded table metadata.",
        ) from exc

    if (
        metadata
        and metadata.constructed_table_id
        and dialect_name == "databricks"
    ):
        try:
            _drop_table_if_exists(engine, qualified_table)
            logger.info("upload:create-table:staging-dropped name=%s", sanitized_table_name)
        except Exception:  # pragma: no cover - best-effort cleanup for warehouse staging tables
            logger.warning("Failed to drop staging table '%s' after replication", qualified_table)

    return UploadDataCreateResponse(
        table_name=sanitized_table_name,
        schema_name=resolved_schema,
        catalog=effective_catalog if dialect_name == "databricks" else None,
        rows_inserted=len(rows),
        target_warehouse=target,
        table_id=metadata.table_id if metadata else None,
        constructed_table_id=metadata.constructed_table_id if metadata else None,
        data_definition_id=metadata.data_definition_id if metadata else None,
        data_definition_table_id=metadata.data_definition_table_id if metadata else None,
    )


@router.delete("/tables/{table_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_uploaded_table(table_id: UUID, db: Session = Depends(get_db)) -> Response:
    table = (
        db.query(Table)
        .options(
            selectinload(Table.definition_tables)
            .selectinload(DataDefinitionTable.constructed_table)
            .selectinload(ConstructedTable.fields)
        )
        .filter(Table.id == table_id)
        .one_or_none()
    )
    if table is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uploaded table was not found.")

    is_constructed = (table.table_type or "").lower() == "constructed" or any(
        definition_table.is_construction for definition_table in table.definition_tables
    )
    if not is_constructed:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uploaded table was not found.")

    engine = get_ingestion_engine()
    dialect_name = engine.dialect.name
    physical_name = table.physical_name or table.name
    if physical_name:
        qualified_table = _build_qualified_table_name(
            physical_name,
            table.schema_name,
            None,
            dialect_name,
        )
        _drop_table_if_exists(engine, qualified_table)

    constructed_tables = [definition_table.constructed_table for definition_table in table.definition_tables if definition_table.constructed_table]

    store = ConstructedDataStore()
    warehouse = None
    try:
        warehouse = ConstructedDataWarehouse()
    except Exception as exc:  # pragma: no cover - defensive fallback
        logger.warning("Unable to initialize constructed data warehouse during delete for table %s: %s", table_id, exc)
        warehouse = None

    for constructed_table in constructed_tables:
        _ = list(constructed_table.fields)
        store.delete_rows_for_table(constructed_table.id)
        if warehouse:
            try:
                warehouse.drop_table(constructed_table)
            except ConstructedDataWarehouseError as exc:
                logger.warning(
                    "Failed to drop constructed data warehouse table '%s': %s",
                    constructed_table.name,
                    exc,
                )
        db.delete(constructed_table)

    for definition_table in list(table.definition_tables):
        db.delete(definition_table)

    db.delete(table)

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.exception("Failed to delete uploaded table %s", table_id)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Unable to delete uploaded table.",
        ) from exc

    return Response(status_code=status.HTTP_204_NO_CONTENT)


async def _read_upload_bytes(upload: UploadFile) -> bytes:
    data = await upload.read()
    if not data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty.")
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file exceeds the 5 MB size limit.",
        )
    return data


def _resolve_delimiter(filename: Optional[str], override: Optional[str]) -> str:
    if override:
        if len(override) != 1:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Delimiter must be a single character.")
        return override
    if filename:
        extension = Path(filename).suffix.lower()
        if extension in {".tsv", ".tab"}:
            return "\t"
    return ","


def _decode_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unable to decode file. Use UTF-8 encoding.")


def _parse_uploaded_table_data(
    *,
    data: bytes,
    filename: Optional[str],
    has_header: bool,
    delimiter: Optional[str],
) -> tuple[list[UploadDataColumn], list[list[Optional[str]]]]:
    if _is_excel_file(filename):
        return _parse_excel_data(data, has_header=has_header)

    detected_delimiter = _resolve_delimiter(filename, delimiter)
    return _parse_csv_data(data, delimiter=detected_delimiter, has_header=has_header)


def _is_excel_file(filename: Optional[str]) -> bool:
    if not filename:
        return False
    suffix = Path(filename).suffix.lower()
    return suffix in {".xlsx", ".xlsm"}


def _parse_csv_data(
    data: bytes,
    *,
    delimiter: str,
    has_header: bool,
) -> tuple[list[UploadDataColumn], list[list[Optional[str]]]]:
    text_data = _decode_bytes(data)
    stream = io.StringIO(text_data)
    reader = csv.reader(stream, delimiter=delimiter)

    raw_rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not raw_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file contains no data.")

    column_count = max(len(row) for row in raw_rows)
    if column_count == 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file has no columns.")

    if has_header:
        header_row = raw_rows[0]
        data_rows = raw_rows[1:]
    else:
        header_row = [f"column_{index + 1}" for index in range(column_count)]
        data_rows = raw_rows

    if len(header_row) < column_count:
        header_row = header_row + [f"column_{index + 1}" for index in range(len(header_row), column_count)]

    normalized_rows = [_normalize_row(row, column_count) for row in data_rows]
    columns = _build_columns(header_row, normalized_rows, column_count)
    return columns, normalized_rows


def _parse_excel_data(
    data: bytes,
    *,
    has_header: bool,
) -> tuple[list[UploadDataColumn], list[list[Optional[str]]]]:
    try:
        workbook = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded Excel file is invalid.",
        ) from exc
    except Exception as exc:  # pragma: no cover - defensive fallback for corrupted files
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unable to read the uploaded Excel file.",
        ) from exc

    try:
        sheet = workbook.active
        raw_rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            formatted = _format_excel_row(row)
            if formatted:
                raw_rows.append(formatted)
    finally:
        workbook.close()

    if not raw_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file contains no data.")

    column_count = max(len(row) for row in raw_rows)
    if column_count == 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file has no columns.")

    if has_header:
        header_row = raw_rows[0]
        data_rows = raw_rows[1:]
    else:
        header_row = [f"column_{index + 1}" for index in range(column_count)]
        data_rows = raw_rows

    if len(header_row) < column_count:
        header_row = header_row + [f"column_{index + 1}" for index in range(len(header_row), column_count)]

    normalized_rows = [_normalize_row(row, column_count) for row in data_rows]
    columns = _build_columns(header_row, normalized_rows, column_count)
    return columns, normalized_rows


def _format_excel_row(row: Sequence[Any]) -> list[str]:
    formatted = [_stringify_excel_cell(cell) for cell in row]
    while formatted and not formatted[-1].strip():
        formatted.pop()
    if not formatted:
        return []
    if not any(cell.strip() for cell in formatted):
        return []
    return formatted


def _stringify_excel_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def _parse_column_overrides(payload: str) -> list[UploadDataColumnOverride]:
    try:
        raw_data = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Column overrides payload must be valid JSON.",
        ) from exc

    if not isinstance(raw_data, list):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Column overrides payload must be a list.",
        )

    overrides: list[UploadDataColumnOverride] = []
    for entry in raw_data:
        if not isinstance(entry, dict):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Column override entries must be objects.",
            )
        try:
            overrides.append(UploadDataColumnOverride(**entry))
        except ValidationError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid column override entry provided.",
            ) from exc
    return overrides


def _normalize_row(row: Sequence[str], column_count: int) -> list[Optional[str]]:
    normalized: list[Optional[str]] = []
    for index in range(column_count):
        value = row[index] if index < len(row) else ""
        normalized.append(_normalize_cell(value))
    return normalized


def _normalize_cell(value: str) -> Optional[str]:
    trimmed = value.strip()
    if not trimmed:
        return None
    lowered = trimmed.lower()
    if lowered in {"null", "none", "nan"}:
        return None
    return trimmed


def _build_columns(
    headers: Sequence[str],
    rows: Sequence[Sequence[Optional[str]]],
    column_count: int,
) -> list[UploadDataColumn]:
    seen: set[str] = set()
    columns: list[UploadDataColumn] = []
    for index in range(column_count):
        original = headers[index] if index < len(headers) else f"column_{index + 1}"
        field_name = _sanitize_identifier(original, f"column_{index + 1}", seen)
        column_values = [row[index] if index < len(row) else None for row in rows]
        inferred_type = _infer_column_type(column_values)
        columns.append(
            UploadDataColumn(
                original_name=(original or "").strip() or field_name,
                field_name=field_name,
                inferred_type=inferred_type,
            )
        )
    return columns


def _sanitize_identifier(source: str, prefix: str, seen: set[str]) -> str:
    candidate = (source or "").strip()
    candidate = IDENTIFIER_SANITIZE_PATTERN.sub("_", candidate)
    candidate = re.sub(r"_+", "_", candidate).strip("_")
    if not candidate:
        candidate = prefix
    if candidate[0].isdigit():
        candidate = f"{prefix}_{candidate}"
    candidate = candidate.lower()
    candidate = candidate[:128]
    base = candidate
    suffix = 1
    while candidate in seen:
        candidate = f"{base}_{suffix}"
        suffix += 1
    seen.add(candidate)
    return candidate


def _sanitize_table_name(name: str) -> str:
    candidate = IDENTIFIER_SANITIZE_PATTERN.sub("_", (name or "").strip())
    candidate = re.sub(r"_+", "_", candidate).strip("_")
    if not candidate:
        candidate = "uploaded_table"
    if candidate[0].isdigit():
        candidate = f"tbl_{candidate}"
    candidate = candidate.lower()
    return candidate[:128]


def _infer_column_type(values: Iterable[Optional[str]]) -> str:
    filtered = [value.strip() for value in values if isinstance(value, str) and value.strip()]
    if not filtered:
        return "string"
    if all(_is_boolean(value) for value in filtered):
        return "boolean"
    if all(_is_integer(value) for value in filtered):
        return "integer"
    if all(_is_float(value) for value in filtered):
        return "float"
    if all(_is_datetime(value) for value in filtered):
        return "timestamp"
    return "string"


def _is_boolean(value: str) -> bool:
    lowered = value.lower()
    return lowered in TRUE_VALUES or lowered in FALSE_VALUES


def _is_integer(value: str) -> bool:
    return bool(re.fullmatch(r"[-+]?\d+", value.strip()))


def _is_float(value: str) -> bool:
    try:
        float(value)
    except ValueError:
        return False
    return True


def _is_datetime(value: str) -> bool:
    candidate = value.strip()
    if candidate.endswith("Z"):
        candidate = candidate[:-1] + "+00:00"
    try:
        datetime.fromisoformat(candidate)
    except ValueError:
        return False
    return True


def _build_qualified_table_name(
    table_name: str,
    schema_name: Optional[str],
    catalog: Optional[str],
    dialect_name: str,
) -> str:
    parts: list[str] = []
    if catalog:
        parts.append(_quote_identifier(catalog, dialect_name))
    if schema_name:
        parts.append(_quote_identifier(schema_name, dialect_name))
    parts.append(_quote_identifier(table_name, dialect_name))
    return ".".join(parts)


def _quote_identifier(identifier: str, dialect_name: str) -> str:
    quote_char = "`" if dialect_name == "databricks" else '"'
    escaped = identifier.replace(quote_char, quote_char * 2)
    return f"{quote_char}{escaped}{quote_char}"


def _apply_column_overrides(
    columns: Sequence[UploadDataColumn],
    overrides: Sequence[UploadDataColumnOverride],
) -> tuple[list[UploadDataColumn], list[int]]:
    if not overrides:
        indices = list(range(len(columns)))
        return list(columns), indices

    valid_fields = {column.field_name.lower() for column in columns}
    override_map: dict[str, UploadDataColumnOverride] = {}
    for override in overrides:
        key = override.field_name.lower()
        if key not in valid_fields:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Column override references unknown field '{override.field_name}'.",
            )
        if key in override_map:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Duplicate column override provided for '{override.field_name}'.",
            )
        override_map[key] = override

    seen: set[str] = set()
    updated: list[UploadDataColumn] = []
    included_indexes: list[int] = []
    for index, column in enumerate(columns):
        override = override_map.get(column.field_name.lower())
        if override and override.exclude is True:
            continue
        fallback_name = f"column_{index + 1}"
        name_source = override.target_name if override and override.target_name else column.field_name
        sanitized_name = _sanitize_override_name(name_source, fallback_name)
        normalized_name = sanitized_name.lower()
        if normalized_name in seen:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Duplicate column names detected. Adjust the column names to be unique.",
            )
        if override and override.target_type:
            column.inferred_type = _normalize_override_type(override.target_type)
        column.field_name = sanitized_name
        seen.add(normalized_name)
        included_indexes.append(index)
        updated.append(column)

    if not updated:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="At least one column must be included for upload.",
        )

    return updated, included_indexes


def _filter_row_values(
    rows: Sequence[Sequence[Optional[str]]],
    indexes: Sequence[int],
) -> list[list[Optional[str]]]:
    if not indexes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No columns selected for upload.")
    filtered: list[list[Optional[str]]] = []
    for row in rows:
        filtered.append([row[index] if index < len(row) else None for index in indexes])
    return filtered


def _sanitize_override_name(source: str, fallback: str) -> str:
    candidate = IDENTIFIER_SANITIZE_PATTERN.sub("_", (source or "").strip())
    candidate = re.sub(r"_+", "_", candidate).strip("_")
    if not candidate:
        candidate = fallback
    if candidate[0].isdigit():
        candidate = f"{fallback}_{candidate}"
    candidate = candidate.lower()
    return candidate[:128]


def _normalize_override_type(value: str) -> str:
    normalized = (value or "").strip().lower()
    allowed = {"boolean", "integer", "float", "timestamp", "string"}
    if normalized not in allowed:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported column type override '{value}'.",
        )
    return normalized


def _apply_create_table(
    connection: Connection,
    qualified_table: str,
    table_name: str,
    schema_name: Optional[str],
    columns: Sequence[UploadDataColumn],
    rows: Sequence[Sequence[Optional[str]]],
    mode: UploadTableMode,
    target: DataWarehouseTarget,
    dialect_name: str,
) -> None:
    if not columns:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No columns detected in uploaded file.")

    if mode is UploadTableMode.REPLACE:
        connection.execute(text(f"DROP TABLE IF EXISTS {qualified_table}"))
    elif mode is not UploadTableMode.CREATE:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported upload mode.")

    spark_loader: SparkTableLoader | None = None
    if (
        target is DataWarehouseTarget.DATABRICKS_SQL
        and dialect_name == "databricks"
        and rows
    ):
        spark_loader = _initialize_spark_loader()

    if spark_loader is not None:
        if mode is UploadTableMode.CREATE and _table_exists(connection, qualified_table):
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Table already exists.")
        try:
            _load_table_with_spark(
                loader=spark_loader,
                schema_name=schema_name,
                table_name=table_name,
                columns=columns,
                rows=rows,
                mode=mode,
            )
            return
        except Exception as exc:  # pragma: no cover - defensive fallback
            logger.warning("Spark load failed for %s; falling back to SQL inserts", qualified_table, exc_info=exc)
            connection.execute(text(f"DROP TABLE IF EXISTS {qualified_table}"))

    column_definitions = ", ".join(
        f"{_quote_identifier(column.field_name, dialect_name)} {_resolve_column_sql_type(column.inferred_type, dialect_name, target)}"
        for column in columns
    )

    connection.execute(text(f"CREATE TABLE {qualified_table} ({column_definitions})"))

    if rows:
        _bulk_insert(connection, qualified_table, columns, rows, dialect_name)


def _resolve_default_schema(
    target: DataWarehouseTarget,
    provided_schema: Optional[str],
    dialect_name: str,
) -> Optional[str]:
    if isinstance(provided_schema, str):
        normalized = provided_schema.strip()
        if normalized:
            return normalized

    if target is not DataWarehouseTarget.DATABRICKS_SQL:
        return None

    if dialect_name != "databricks":
        return None

    try:
        params = get_ingestion_connection_params()
    except Exception:  # pragma: no cover - defensive fallback when warehouse not configured
        return ConstructedDataWarehouse.DEFAULT_SCHEMA

    return params.constructed_schema or params.schema_name or ConstructedDataWarehouse.DEFAULT_SCHEMA


def _register_uploaded_table_metadata(
    *,
    table_name: str,
    schema_name: Optional[str],
    columns: Sequence[UploadDataColumn],
    product_team_id: UUID,
    data_object_id: UUID,
    system_id: UUID,
    session: Session,
) -> _UploadMetadataResult:
    process_area = session.get(ProcessArea, product_team_id)
    if not process_area:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selected process area was not found.")

    data_object = session.get(DataObject, data_object_id)
    if not data_object:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selected data object was not found.")
    if data_object.process_area_id != product_team_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Data object does not belong to the selected process area.",
        )

    system = session.get(System, system_id)
    if not system:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selected system was not found.")

    link_exists = (
        session.query(DataObjectSystem)
        .filter(
            DataObjectSystem.data_object_id == data_object_id,
            DataObjectSystem.system_id == system_id,
        )
        .first()
    )
    if not link_exists:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Data object is not linked to the selected system.",
        )

    definition = (
        session.query(DataDefinition)
        .filter(
            DataDefinition.data_object_id == data_object_id,
            DataDefinition.system_id == system_id,
        )
        .one_or_none()
    )
    if not definition:
        definition = DataDefinition(
            data_object_id=data_object_id,
            system_id=system_id,
        )
        session.add(definition)
        session.flush()

    normalized_name = table_name.lower()
    table_query = session.query(Table).filter(
        Table.system_id == system_id,
        func.lower(Table.physical_name) == normalized_name,
    )
    if schema_name:
        table = table_query.filter(func.lower(Table.schema_name) == schema_name.lower()).one_or_none()
    else:
        table = table_query.filter(Table.schema_name.is_(None)).one_or_none()
    if not table:
        table = table_query.first()

    if not table:
        table = Table(
            system_id=system_id,
            name=table_name,
            physical_name=table_name,
            schema_name=schema_name,
            table_type="constructed",
            status="active",
        )
        session.add(table)
        session.flush()
    else:
        table.name = table_name
        table.physical_name = table_name
        table.schema_name = schema_name
        table.table_type = table.table_type or "constructed"
        table.status = table.status or "active"

    desired_names = {column.field_name.lower() for column in columns}
    legacy_audit_names = {"project", "release"}
    existing_fields = session.query(Field).filter(Field.table_id == table.id).all()
    for field in existing_fields:
        field_name = (field.name or "").lower()
        if field_name in legacy_audit_names:
            session.delete(field)
            continue
        if field_name not in desired_names and field_name not in AUDIT_FIELD_NAME_SET:
            session.delete(field)
    session.flush()

    existing_fields_by_name = {
        (field.name or "").lower(): field
        for field in session.query(Field).filter(Field.table_id == table.id).all()
    }
    column_fields: list[Field] = []
    for column in columns:
        key = column.field_name.lower()
        inferred_type = _map_inferred_type_to_field_type(column.inferred_type)
        field = existing_fields_by_name.get(key)
        if field:
            field.field_type = inferred_type
            field.active = True
            if field.system_required:
                field.system_required = False
            if field.business_process_required:
                field.business_process_required = False
        else:
            field = Field(
                table_id=table.id,
                name=column.field_name,
                field_type=inferred_type,
                system_required=False,
                business_process_required=False,
                suppressed_field=False,
                active=True,
            )
            session.add(field)
            session.flush()
            existing_fields_by_name[key] = field
        if field not in table.fields:
            table.fields.append(field)
        column_fields.append(field)

    definition_table = (
        session.query(DataDefinitionTable)
        .filter(
            DataDefinitionTable.data_definition_id == definition.id,
            DataDefinitionTable.table_id == table.id,
        )
        .one_or_none()
    )
    if not definition_table:
        definition_table = DataDefinitionTable(
            data_definition_id=definition.id,
            table_id=table.id,
            alias=table.name,
            is_construction=True,
        )
        session.add(definition_table)
        session.flush()
    else:
        definition_table.is_construction = True
        if not definition_table.alias:
            definition_table.alias = table.name

    definition_table.table = table

    for definition_field in list(definition_table.fields):
        field = definition_field.field
        if not field:
            continue
        field_name = (field.name or "").lower()
        if field_name in legacy_audit_names:
            if definition_field in definition_table.fields:
                definition_table.fields.remove(definition_field)
            session.delete(definition_field)

    existing_definition_fields = {
        field.field_id: field
        for field in session.query(DataDefinitionField)
        .filter(DataDefinitionField.definition_table_id == definition_table.id)
        .all()
    }

    for index, field in enumerate(column_fields):
        definition_field = existing_definition_fields.get(field.id)
        if definition_field:
            definition_field.display_order = index
            definition_field.is_unique = False
        else:
            definition_field = DataDefinitionField(
                definition_table_id=definition_table.id,
                field_id=field.id,
                display_order=index,
                is_unique=False,
            )
            session.add(definition_field)
        if definition_field not in definition_table.fields:
            definition_table.fields.append(definition_field)

    session.flush()
    _ensure_audit_fields_for_definition_table(definition_table, session)
    session.flush()

    sync_construction_tables_for_definition(definition.id, session)
    session.flush()

    constructed_table = (
        session.query(ConstructedTable)
        .filter(ConstructedTable.data_definition_table_id == definition_table.id)
        .one_or_none()
    )
    constructed_table_id = constructed_table.id if constructed_table else None

    return _UploadMetadataResult(
        table_id=table.id,
        data_definition_id=definition.id,
        data_definition_table_id=definition_table.id,
        constructed_table_id=constructed_table_id,
    )


def _sync_constructed_table_rows(
    *,
    constructed_table_id: UUID,
    columns: Sequence[UploadDataColumn],
    rows: Sequence[Sequence[Optional[str]]],
    session: Session,
    force_recreate_warehouse: bool = False,
) -> None:
    if not rows:
        return

    constructed_table = (
        session.query(ConstructedTable)
        .options(selectinload(ConstructedTable.fields))
        .filter(ConstructedTable.id == constructed_table_id)
        .one_or_none()
    )
    if constructed_table is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Constructed table metadata could not be loaded.",
        )

    # Ensure relationship data is loaded before any external operations.
    _ = list(constructed_table.fields)

    audit_defaults = _build_audit_defaults(constructed_table, session)

    store = ConstructedDataStore()
    store.delete_rows_for_table(constructed_table_id)

    payloads = [_build_payload_from_row(columns, row, audit_defaults) for row in rows]
    records = store.insert_rows(constructed_table_id, payloads)

    try:
        warehouse = ConstructedDataWarehouse()
        if force_recreate_warehouse:
            warehouse.drop_table(constructed_table)
        warehouse.ensure_table(constructed_table)
        if records:
            warehouse.upsert_records(constructed_table, records)
    except ConstructedDataWarehouseError as exc:
        store.delete_rows_for_table(constructed_table_id)
        logger.error("Failed to populate constructed table %s: %s", constructed_table_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Unable to copy uploaded rows into constructed table.",
        ) from exc
    except Exception as exc:  # pragma: no cover - defensive fallback
        store.delete_rows_for_table(constructed_table_id)
        logger.error("Unexpected error populating constructed table %s: %s", constructed_table_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Unable to copy uploaded rows into constructed table.",
        ) from exc


def _build_payload_from_row(
    columns: Sequence[UploadDataColumn],
    row: Sequence[Optional[str]],
    audit_defaults: dict[str, Optional[str]],
) -> dict[str, Optional[str]]:
    payload: dict[str, Optional[str]] = {}
    for index, column in enumerate(columns):
        payload[column.field_name] = row[index] if index < len(row) else None
    for key, value in audit_defaults.items():
        if key not in payload and value is not None:
            payload[key] = value
    return payload


def _build_audit_defaults(constructed_table: ConstructedTable, session: Session) -> dict[str, Optional[str]]:
    _ = (constructed_table, session)  # parameters retained for future use

    now = datetime.now(timezone.utc).isoformat()
    return {
        "Created By": "Upload Data",
        "Created Date": now,
        "Modified By": "Upload Data",
        "Modified Date": now,
    }


def _map_inferred_type_to_field_type(inferred: str) -> str:
    mapping = {
        "boolean": "boolean",
        "integer": "integer",
        "float": "float",
        "timestamp": "timestamp",
        "string": "string",
    }
    return mapping.get((inferred or "").lower(), "string")


def _drop_table_if_exists(engine: Engine, qualified_table: str) -> None:
    try:
        with engine.begin() as connection:
            connection.execute(text(f"DROP TABLE IF EXISTS {qualified_table}"))
    except SQLAlchemyError:  # pragma: no cover - best-effort cleanup
        return


def _table_exists(connection: Connection, qualified_table: str) -> bool:
    try:
        connection.execute(text(f"SELECT 1 FROM {qualified_table} LIMIT 1"))
    except SQLAlchemyError:
        return False
    return True


def _initialize_spark_loader() -> SparkTableLoader | None:
    try:
        return SparkTableLoader()
    except RuntimeError as exc:
        logger.info("Spark loader unavailable: %s", exc)
    except Exception as exc:  # pragma: no cover - defensive fallback
        logger.warning("Unable to initialize Spark loader: %s", exc)
    return None


def _load_table_with_spark(
    *,
    loader: SparkTableLoader,
    schema_name: Optional[str],
    table_name: str,
    columns: Sequence[UploadDataColumn],
    rows: Sequence[Sequence[Optional[str]]],
    mode: UploadTableMode,
) -> None:
    typed_rows = _build_spark_rows(columns, rows)
    if not typed_rows:
        return

    plan = build_loader_plan(
        schema=schema_name,
        table_name=table_name,
        replace=(mode is UploadTableMode.REPLACE),
        deduplicate=False,
    )
    loader_columns = _build_sqlalchemy_columns(columns)
    loader.load_rows(plan, typed_rows, loader_columns)


def _build_spark_rows(
    columns: Sequence[UploadDataColumn],
    rows: Sequence[Sequence[Optional[str]]],
) -> list[Mapping[str, object | None]]:
    payloads: list[Mapping[str, object | None]] = []
    for row in rows:
        record: dict[str, object | None] = {}
        for index, column in enumerate(columns):
            raw_value = row[index] if index < len(row) else None
            record[column.field_name] = _coerce_row_value(raw_value, column.inferred_type)
        payloads.append(record)
    return payloads


def _build_sqlalchemy_columns(columns: Sequence[UploadDataColumn]) -> list[Column]:
    loader_columns: list[Column] = []
    for column in columns:
        loader_columns.append(Column(column.field_name, _map_inferred_type_to_sqlalchemy(column.inferred_type), nullable=True))
    return loader_columns


def _coerce_row_value(value: Optional[str], inferred_type: str) -> object | None:
    if value is None:
        return None
    candidate = value.strip()
    if not candidate:
        return None

    lowered = candidate.lower()
    if inferred_type == "boolean":
        if lowered in TRUE_VALUES:
            return True
        if lowered in FALSE_VALUES:
            return False
        return None
    if inferred_type == "integer":
        try:
            return int(candidate)
        except ValueError:
            return None
    if inferred_type == "float":
        try:
            return float(candidate)
        except ValueError:
            return None
    if inferred_type == "timestamp":
        normalized = candidate[:-1] + "+00:00" if candidate.endswith("Z") else candidate
        try:
            return datetime.fromisoformat(normalized)
        except ValueError:
            return None
    return candidate


def _map_inferred_type_to_sqlalchemy(column_type: str) -> sqltypes.TypeEngine:
    mapping = {
        "boolean": sqltypes.Boolean(),
        "integer": sqltypes.BigInteger(),
        "float": sqltypes.Float(),
        "timestamp": sqltypes.DateTime(timezone=True),
        "string": sqltypes.String(),
    }
    return mapping.get((column_type or "").lower(), sqltypes.String())


def _resolve_column_sql_type(column_type: str, dialect_name: str, target: DataWarehouseTarget) -> str:
    if dialect_name == "sqlite":
        mapping = {
            "boolean": "INTEGER",
            "integer": "INTEGER",
            "float": "REAL",
            "timestamp": "TEXT",
            "string": "TEXT",
        }
        return mapping.get(column_type, "TEXT")

    if target is DataWarehouseTarget.DATABRICKS_SQL:
        mapping = {
            "boolean": "BOOLEAN",
            "integer": "BIGINT",
            "float": "DOUBLE",
            "timestamp": "TIMESTAMP",
            "string": "STRING",
        }
        return mapping.get(column_type, "STRING")

    mapping = {
        "boolean": "BOOLEAN",
        "integer": "BIGINT",
        "float": "DOUBLE PRECISION",
        "timestamp": "TIMESTAMP",
        "string": "TEXT",
    }
    return mapping.get(column_type, "TEXT")


def _bulk_insert(
    connection: Connection,
    qualified_table: str,
    columns: Sequence[UploadDataColumn],
    rows: Sequence[Sequence[Optional[str]]],
    dialect_name: str,
) -> None:
    column_list = ", ".join(_quote_identifier(column.field_name, dialect_name) for column in columns)
    chunk_size = 500
    column_count = len(columns)

    for offset in range(0, len(rows), chunk_size):
        chunk = rows[offset : offset + chunk_size]
        if not chunk:
            continue
        values_sql = ", ".join(
            "("
            + ", ".join(
                _format_value(row[index] if index < len(row) else None, columns[index].inferred_type, dialect_name)
                for index in range(column_count)
            )
            + ")"
            for row in chunk
        )
        statement = f"INSERT INTO {qualified_table} ({column_list}) VALUES {values_sql}"
        connection.execute(text(statement))


def _format_value(value: Optional[str], column_type: str, dialect_name: str) -> str:
    if value is None:
        return "NULL"

    if column_type == "boolean":
        lowered = value.lower()
        if lowered in TRUE_VALUES:
            return "1" if dialect_name == "sqlite" else "TRUE"
        if lowered in FALSE_VALUES:
            return "0" if dialect_name == "sqlite" else "FALSE"
        return "NULL"

    if column_type == "integer":
        try:
            int(value)
        except ValueError:
            return "NULL"
        return value.strip()

    if column_type == "float":
        try:
            float(value)
        except ValueError:
            return "NULL"
        return value.strip()

    escaped = value.replace("'", "''")
    return f"'{escaped}'"
